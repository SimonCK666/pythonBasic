#
# xslSpsheet.py
# @author bulbasaur
# @description Reduce the price 10% and add beautiful charts behind the excel
# @created 2020-03-08T19:54:49.822Z+08:00
# @last-modified 2020-03-08T23:03:10.093Z+08:00
#

import openpyxl as xl

myfile = open("E:\python_basic\python_6hours\ExcleSheet\sheets.xlsx", "r")
wb = xl.load_workbook('E:\python_basic\python_6hours\ExcleSheet\sheets.xlsx')
# print(wb.sheetnames)

sheet = wb['Sheet1']

# find the special cell
# cell = sheet['a1']
cell = sheet.cell(1, 1)
print(cell.value)
